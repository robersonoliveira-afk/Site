#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simulador de Projeto Agropecuário — Versão Excel
Gera Simulador_Agropecuario.xlsx com 8 abas, fórmulas interligadas,
dropdowns de validação e formatação visual completa.

Dependência: pip install openpyxl
Uso: python simulador_agro_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

OUTPUT_FILE = "Simulador_Agropecuario.xlsx"
C_TITULO = "2E5C3F"; C_ENTRADA = "FFF9C4"; C_CALC = "E3F2FD"
C_ALERTA = "FFCDD2"; C_OK = "C8E6C9"; C_DESTAQUE = "FFF3E0"; C_CINZA = "F5F5F5"
THIN = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
              top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

def estilo(cell, tipo="normal", negrito=False, tamanho=10, cor_fonte="000000"):
    if tipo == "titulo":
        cell.fill = PatternFill(start_color=C_TITULO, end_color=C_TITULO, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=tamanho, name="Calibri")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    elif tipo == "entrada":
        cell.fill = PatternFill(start_color=C_ENTRADA, end_color=C_ENTRADA, fill_type="solid")
        cell.font = Font(color=cor_fonte, bold=negrito, size=tamanho, name="Calibri")
        cell.border = THIN; cell.protection = Protection(locked=False)
    elif tipo == "calculado":
        cell.fill = PatternFill(start_color=C_CALC, end_color=C_CALC, fill_type="solid")
        cell.font = Font(color="0D47A1", bold=negrito, size=tamanho, name="Calibri")
        cell.border = THIN; cell.protection = Protection(locked=True)
    elif tipo == "alerta":
        cell.fill = PatternFill(start_color=C_ALERTA, end_color=C_ALERTA, fill_type="solid")
        cell.font = Font(color="B71C1C", bold=True, size=tamanho, name="Calibri")
    elif tipo == "ok":
        cell.fill = PatternFill(start_color=C_OK, end_color=C_OK, fill_type="solid")
        cell.font = Font(color="1B5E20", bold=True, size=tamanho, name="Calibri")
    elif tipo == "destaque":
        cell.fill = PatternFill(start_color=C_DESTAQUE, end_color=C_DESTAQUE, fill_type="solid")
        cell.font = Font(color="E65100", bold=negrito, size=tamanho, name="Calibri")
        cell.border = THIN
    else:
        cell.font = Font(color=cor_fonte, bold=negrito, size=tamanho, name="Calibri")
        cell.border = THIN
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def auto_largura(ws, min_w=8, max_w=55):
    for col in ws.columns:
        first = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                first = cell; break
        if first is None: continue
        col_letter = first.column_letter; max_len = 0
        for cell in col:
            if not hasattr(cell, 'column_letter'): continue
            try:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), max_w)

def nomear(wb, nome, aba, celula):
    dn = DefinedName(nome, attr_text=f"'{aba}'!{celula}")
    wb.defined_names.add(dn)

def titulo_secao(ws, linha, texto, col_ini="B", col_fim="H"):
    ws.merge_cells(f"{col_ini}{linha}:{col_fim}{linha}")
    cell = ws[f"{col_ini}{linha}"]
    cell.value = texto; estilo(cell, "titulo", tamanho=11)
    ws.row_dimensions[linha].height = 22

def rotulo(ws, linha, texto, col="B"):
    cell = ws[f"{col}{linha}"]; cell.value = texto
    estilo(cell, "entrada", negrito=True)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    return cell

def entrada(ws, linha, col="C", valor="", fmt="General"):
    cell = ws[f"{col}{linha}"]; cell.value = valor; estilo(cell, "entrada")
    if fmt != "General": cell.number_format = fmt
    return cell

def calc(ws, linha, col, formula, fmt="General", negrito=False, tamanho=10):
    cell = ws[f"{col}{linha}"]; cell.value = formula; estilo(cell, "calculado", negrito=negrito, tamanho=tamanho)
    if fmt != "General": cell.number_format = fmt
    return cell

TEMPLATES = [
    ["Codigo","Nome","Unidade","LblQtd","LblProd","UnProd","UnPreco","Familia","Modo","Perene","Giro"],
    ["graos","Grãos e cereais","ha","Área","Produtividade","sc/ha","R$/sc","lavoura","A",0,0],
    ["fumo","Fumo (tabaco)","ha","Área","Produtividade","@/ha","R$/@","lavoura","A",0,0],
    ["horta","Horta / olericultura","ha","Área","Produtividade","kg/ha","R$/kg","lavoura","A",0,0],
    ["fruta","Fruticultura","ha","Área","Produtividade","kg/ha","R$/kg","lavoura","B",1,0],
    ["ervamate","Erva-mate","ha","Área","Produtividade","kg/ha","R$/kg","lavoura","B",1,0],
    ["madeira","Silvicultura","ha","Área","Produtividade","m³/ha","R$/m³","lavoura","H",1,0],
    ["corte","Corte / confinamento","cabeça","Cabeças (giro)","Peso de venda","kg PV/cab","R$/kg vivo","criacao","D",0,1],
    ["leite","Leite","vaca","Vacas em lactação","Produção","L/vaca/ano","R$/L","criacao","C",0,0],
    ["cria","Cria (bovinos)","matriz","Matrizes","Produção","kg terneiro/matriz/ano","R$/kg","criacao","C",0,0],
    ["ovinos","Ovinos","matriz","Matrizes (ovelhas)","Produção","kg PV/matriz/ano","R$/kg PV","criacao","C",0,0],
    ["suinos","Suínos","cabeça","Cabeças alojadas","Peso de abate","kg/cab","R$/kg","criacao","D",0,1],
    ["aves_corte","Aves de corte","cabeça","Cabeças alojadas","Peso de abate","kg/cab","R$/kg","criacao","D",0,1],
    ["aves_postura","Galinhas poedeiras","matriz","Matrizes (galinhas)","Produção","dúzias/matriz/ano","R$/dúzia","postura","G",0,0],
    ["apicultura","Apicultura","colmeia","Colmeias","Produção","kg mel/colmeia/ano","R$/kg","apicultura","F",0,0],
    ["tilapia","Tilápia","tanque","Tanques","Produção","kg/tanque","R$/kg","criacao","D",0,1],
    ["outra","Outra","unidade","Unidades","Produtividade","un.","R$/un.","outra","E",0,0],
]
CATEGORIAS = [
    ["Familia","Categoria"],["lavoura","Sementes/mudas"],["lavoura","Fertilizantes"],
    ["lavoura","Corretivos"],["lavoura","Agrotóxicos"],["lavoura","Operações/mecanização"],
    ["lavoura","Mão de obra"],["lavoura","Outros"],["criacao","Alimentação"],
    ["criacao","Sanidade"],["criacao","Aquisição de animais (giro)"],["criacao","Mão de obra"],
    ["criacao","Energia/água"],["criacao","Outros"],["apicultura","Alimentação suplementar"],
    ["apicultura","Sanidade"],["apicultura","Rainha e enxames"],["apicultura","Materiais e embalagens"],
    ["apicultura","Mão de obra"],["apicultura","Outros"],["postura","Ração e suplementos"],
    ["postura","Sanidade"],["postura","Energia e aquecimento"],["postura","Embalagens e materiais"],
    ["postura","Mão de obra"],["postura","Outros"],["outra","Insumos"],["outra","Mão de obra"],
    ["outra","Operações"],["outra","Outros"],
]
CULTURAS = [
    ["Codigo","Nome","Form","Vida"],["uva","Uva",3,20],["maca","Maçã",4,20],
    ["citros","Citros",3,20],["pessego","Pêssego",3,12],["erva","Erva-mate",5,30],
    ["peca","Noz-pecã",7,40],["oliveira","Oliveira",4,30],["eucalipto","Eucalipto",7,7],
]
FUNRURAL = [["Fiscal","Aliquota"],["PF",0.0163],["PJ",0.0223]]

def construir_workbook():
    wb = Workbook()
    # ABA Ref
    ws_ref = wb.active; ws_ref.title = "Ref"; ws_ref.sheet_state = "hidden"
    for r_idx, row in enumerate(TEMPLATES, 1):
        for c_idx, val in enumerate(row, 1): ws_ref.cell(row=r_idx, column=c_idx, value=val)
    for r_idx, row in enumerate(CATEGORIAS, 1):
        for c_idx, val in enumerate(row, 1): ws_ref.cell(row=r_idx, column=c_idx+12, value=val)
    for r_idx, row in enumerate(CULTURAS, 1):
        for c_idx, val in enumerate(row, 1): ws_ref.cell(row=r_idx, column=c_idx+15, value=val)
    for r_idx, row in enumerate(FUNRURAL, 1):
        for c_idx, val in enumerate(row, 1): ws_ref.cell(row=r_idx, column=c_idx+20, value=val)
    all_cats = sorted(set([r[1] for r in CATEGORIAS[1:]]))
    for i, cat in enumerate(all_cats, 2): ws_ref.cell(row=i, column=24, value=cat)

    # ABA 1 — Identificação
    ws = wb.create_sheet("1.Identificacao")
    ws.sheet_properties.tabColor = "2E5C3F"
    titulo_secao(ws, 1, "1. IDENTIFICAÇÃO DO PROJETO", "B", "E")
    for l, v, t, d in [(3,"C3","Título do projeto:",""),(4,"C4","Produtor:",""),
                       (5,"C5","Unidade de produção:",""),(6,"C6","Município:",""),(7,"C7","Data:","")]:
        rotulo(ws, l, t, "B"); entrada(ws, l, v[0], d)
    for n, c in [("projeto_val","C3"),("produtor_val","C4"),("unidade_val","C5"),
                 ("municipio_val","C6"),("data_val","C7")]: nomear(wb, n, "1.Identificacao", c)
    rotulo(ws, 9, "Template (atividade):", "B")
    dv_t = DataValidation(type="list", formula1="=Ref!$A$2:$A$17", allow_blank=False)
    dv_t.error = "Selecione um template válido"; ws.add_data_validation(dv_t)
    c = entrada(ws, 9, "C", "graos"); dv_t.add(c); nomear(wb, "template_val", "1.Identificacao", "C9")
    rotulo(ws, 10, "Enquadramento fiscal:", "B")
    dv_f = DataValidation(type="list", formula1="=Ref!$U$2:$U$3", allow_blank=False)
    ws.add_data_validation(dv_f); c = entrada(ws, 10, "C", "PF"); dv_f.add(c)
    nomear(wb, "fiscal_val", "1.Identificacao", "C10")
    titulo_secao(ws, 12, "CRÉDITO RURAL", "B", "E")
    for l, v, t, lista, d in [(13,"C13","Usa crédito?",["sim","nao"],"nao"),
        (14,"C14","Finalidade:",["Custeio","Investimento","Comercialização","Industrialização"],""),
        (15,"C15","Enquadramento:",["Pronaf","Pronamp","Demais produtores"],""),
        (16,"C16","Documentos organizados?",["sim","nao"],"nao"),
        (17,"C17","Módulos fiscais:",None,""),(18,"C18","Localização:",None,"")]:
        rotulo(ws, l, t, "B"); c = entrada(ws, l, v, d)
        if lista:
            dv = DataValidation(type="list", formula1=f'="{",".join(lista)}"', allow_blank=True)
            ws.add_data_validation(dv); dv.add(c)
    titulo_secao(ws, 20, "DIAGNÓSTICO ESTRATÉGICO (SWOT)", "B", "E")
    for l, v, t in [(21,"C21","Forças:"),(22,"C22","Fraquezas:"),(23,"C23","Oportunidades:"),(24,"C24","Ameaças:")]:
        rotulo(ws, l, t, "B"); c = entrada(ws, l, v, ""); c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[l].height = 50
    auto_largura(ws)

    # ABA 2 — Receita
    ws = wb.create_sheet("2.Receita"); ws.sheet_properties.tabColor = "2E5C3F"
    titulo_secao(ws, 1, "2. PRODUÇÃO E RECEITA", "B", "F")
    rotulo(ws, 3, "Produto / Cultura:", "B"); entrada(ws, 3, "C"); nomear(wb, "produto_nome", "2.Receita", "C3")
    rotulo(ws, 4, "Unidades de produção:", "B"); entrada(ws, 4, "C", 0, "#,##0.00"); nomear(wb, "qtd_val", "2.Receita", "C4")
    rotulo(ws, 4, "Unidade:", "D"); calc(ws, 4, "E", '=VLOOKUP(template_val,Ref!$A$1:$K$17,3,FALSE)', "General")
    rotulo(ws, 5, "Produtividade:", "B"); entrada(ws, 5, "C", 0, "#,##0.00"); nomear(wb, "prod_val", "2.Receita", "C5")
    rotulo(ws, 5, "Unidade:", "D"); calc(ws, 5, "E", '=VLOOKUP(template_val,Ref!$A$1:$K$17,6,FALSE)', "General")
    rotulo(ws, 6, "Preço:", "B"); entrada(ws, 6, "C", 0, 'R$ #,##0.00'); nomear(wb, "preco_val", "2.Receita", "C6")
    rotulo(ws, 6, "Unidade:", "D"); calc(ws, 6, "E", '=VLOOKUP(template_val,Ref!$A$1:$K$17,7,FALSE)', "General")
    titulo_secao(ws, 8, "OUTRAS RECEITAS", "B", "E")
    rotulo(ws, 9, "Descrição:", "B"); entrada(ws, 9, "C"); nomear(wb, "sec_desc_val", "2.Receita", "C9")
    rotulo(ws, 9, "Valor (R$):", "D"); entrada(ws, 9, "E", 0, 'R$ #,##0.00'); nomear(wb, "sec_valor_val", "2.Receita", "E9")
    titulo_secao(ws, 11, "DEDUÇÕES DA RECEITA", "B", "E")
    rotulo(ws, 12, "Comercialização (%):", "B"); entrada(ws, 12, "C", 0, '0.00"%"'); nomear(wb, "comerc_pct", "2.Receita", "C12")
    rotulo(ws, 13, "Funrural (%):", "B")
    calc(ws, 13, "C", '=VLOOKUP(fiscal_val,Ref!$U$1:$V$3,2,FALSE)', '0.00%')
    ws["B15"].value = "RECEITA BRUTA (RB):"; estilo(ws["B15"], "calculado", negrito=True, tamanho=11)
    ws["B15"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 15, "C", '=qtd_val*prod_val*preco_val+sec_valor_val', 'R$ #,##0.00', negrito=True, tamanho=11)
    nomear(wb, "rb_val", "2.Receita", "C15")
    ws["B16"].value = "(−) Comercialização:"; estilo(ws["B16"], "calculado", negrito=True)
    ws["B16"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 16, "C", '=rb_val*(comerc_pct/100)', 'R$ #,##0.00'); nomear(wb, "comerc_rs", "2.Receita", "C16")
    ws["B17"].value = "(−) Funrural:"; estilo(ws["B17"], "calculado", negrito=True)
    ws["B17"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 17, "C", '=rb_val*C13', 'R$ #,##0.00'); nomear(wb, "funrural_rs", "2.Receita", "C17")
    ws["B18"].value = "RECEITA LÍQUIDA (RL):"; estilo(ws["B18"], "ok", negrito=True, tamanho=12)
    ws["B18"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 18, "C", '=rb_val-comerc_rs-funrural_rs', 'R$ #,##0.00', negrito=True, tamanho=12)
    nomear(wb, "rl_val", "2.Receita", "C18")
    ws["B20"].value = "Fórmula: Qtd × Produtividade × Preço + Outras receitas"
    ws["B20"].font = Font(italic=True, color="666666", size=9)
    auto_largura(ws)

    # ABA 3 — Manejo
    ws = wb.create_sheet("3.Manejo"); ws.sheet_properties.tabColor = "2E5C3F"
    titulo_secao(ws, 1, "3. PRÁTICAS DE MANEJO", "B", "E")
    rotulo(ws, 3, "Modo de manejo:", "B")
    calc(ws, 3, "C", '=VLOOKUP(template_val,Ref!$A$1:$K$17,9,FALSE)', "General", negrito=True, tamanho=11)
    nomear(wb, "modo_val", "3.Manejo", "C3")
    for l, v, t in [(5,"C5","Sistema / Espécie / Categoria:"),(6,"C6","Variedade / Cultivar:"),
                    (7,"C7","Mês de início / Plantio:"),(8,"C8","Mês de colheita / Final:"),
                    (9,"C9","Adubação / Dieta / Ração:"),(10,"C10","Sanidade / Tratamentos:"),
                    (11,"C11","Irrigação / Instalações:"),(12,"C12","Outras práticas:")]:
        rotulo(ws, l, t, "B"); entrada(ws, l, v)
    rotulo(ws, 14, "Observações detalhadas:", "B")
    ws.merge_cells("C14:E16"); c = entrada(ws, 14, "C", "")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[14].height = 60
    auto_largura(ws)

    # ABA 4 — COE
    ws = wb.create_sheet("4.COE"); ws.sheet_properties.tabColor = "2E5C3F"
    titulo_secao(ws, 1, "4. CUSTEIO (COE)", "B", "I")
    headers_coe = ["Nº", "Descrição", "Categoria", "Unidade", "Qtd/Un", "Preço Unit (R$)", "Custo/Un (R$)", "Custo Total (R$)"]
    for c_idx, h in enumerate(headers_coe, 2):
        cell = ws.cell(row=3, column=c_idx, value=h); estilo(cell, "titulo", tamanho=10)
    for row in range(4, 24):
        ws.cell(row=row, column=2, value=row-3).alignment = Alignment(horizontal="center", vertical="center")
        for col in [3,4,5]: estilo(ws.cell(row=row, column=col), "entrada")
        dv_cat = DataValidation(type="list", formula1="=Ref!$X$2:$X$30", allow_blank=True)
        ws.add_data_validation(dv_cat); dv_cat.add(ws.cell(row=row, column=4))
        c = ws.cell(row=row, column=6, value=0); estilo(c, "entrada"); c.number_format = '#,##0.00'
        c = ws.cell(row=row, column=7, value=0); estilo(c, "entrada"); c.number_format = 'R$ #,##0.00'
        c = ws.cell(row=row, column=8)
        c.value = f'=IF(AND(F{row}>0,G{row}>0),F{row}*G{row},0)'; estilo(c, "calculado"); c.number_format = 'R$ #,##0.00'
        c = ws.cell(row=row, column=9)
        c.value = f'=IF(H{row}>0,H{row}*qtd_val,0)'; estilo(c, "calculado"); c.number_format = 'R$ #,##0.00'
    ws["B25"].value = "Subtotal dos itens diretos:"; estilo(ws["B25"], "calculado", negrito=True, tamanho=11)
    ws["B25"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 25, "H", '=SUM(I4:I23)', 'R$ #,##0.00', negrito=True, tamanho=11)
    nomear(wb, "subtotal_coe", "4.COE", "H25")
    for l, v, t, n, f in [(27,"C27","Seguro (%):","seguro_pct",'0.00"%"'),
                           (28,"C28","Assistência Técnica (%):","ater_pct",'0.00"%"'),
                           (29,"C29","Juros de custeio (%):","juros_pct",'0.00"%"')]:
        rotulo(ws, l, t, "B"); c = entrada(ws, l, v, 0, f); nomear(wb, n, "4.COE", v)
    calc(ws, 27, "E", '=H25*(seguro_pct/100)', 'R$ #,##0.00'); nomear(wb, "seguro_rs", "4.COE", "E27")
    calc(ws, 28, "E", '=H25*(ater_pct/100)', 'R$ #,##0.00'); nomear(wb, "ater_rs", "4.COE", "E28")
    calc(ws, 29, "E", '=H25*(juros_pct/100)', 'R$ #,##0.00'); nomear(wb, "juros_rs", "4.COE", "E29")
    ws["B31"].value = "COE TOTAL:"; estilo(ws["B31"], "ok", negrito=True, tamanho=12)
    ws["B31"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 31, "H", '=H25+E27+E28+E29', 'R$ #,##0.00', negrito=True, tamanho=12)
    nomear(wb, "coe_total", "4.COE", "H31")
    ws["B33"].value = "Preencha apenas as colunas em AMARELO. As demais calculam automaticamente."
    ws["B33"].font = Font(italic=True, color="666666", size=9)
    auto_largura(ws)

    # ABA 5 — Implantação
    ws = wb.create_sheet("5.Implantacao"); ws.sheet_properties.tabColor = "B89233"
    titulo_secao(ws, 1, "5. IMPLANTAÇÃO (CULTURAS PERENES)", "B", "H")
    rotulo(ws, 3, "É cultura perene?", "B")
    calc(ws, 3, "C", '=IF(VLOOKUP(template_val,Ref!$A$1:$K$17,10,FALSE)=1,"SIM — Preencha esta aba","NÃO — Opcional")', "General", negrito=True, tamanho=11)
    rotulo(ws, 5, "Cultura perene:", "B")
    dv_cult = DataValidation(type="list", formula1="=Ref!$P$2:$P$9", allow_blank=True)
    ws.add_data_validation(dv_cult); c = entrada(ws, 5, "C", "uva"); dv_cult.add(c)
    nomear(wb, "cultura_val", "5.Implantacao", "C5")
    rotulo(ws, 6, "Anos de formação:", "B"); entrada(ws, 6, "C", 0, "0"); nomear(wb, "form_anos", "5.Implantacao", "C6")
    rotulo(ws, 7, "Vida útil produtiva (anos):", "B"); entrada(ws, 7, "C", 0, "0"); nomear(wb, "vida_util", "5.Implantacao", "C7")
    titulo_secao(ws, 9, "ITENS DE IMPLANTAÇÃO (por ha)", "B", "G")
    imp_h = ["Nº", "Descrição", "Unidade", "Qtd/ha", "Preço Unit (R$)", "Custo/ha (R$)"]
    for c_idx, h in enumerate(imp_h, 2):
        cell = ws.cell(row=10, column=c_idx, value=h); estilo(cell, "titulo", tamanho=10)
    for row in range(11, 26):
        ws.cell(row=row, column=2, value=row-10).alignment = Alignment(horizontal="center", vertical="center")
        for col in [3,4]: estilo(ws.cell(row=row, column=col), "entrada")
        c = ws.cell(row=row, column=5, value=0); estilo(c, "entrada"); c.number_format = '#,##0.00'
        c = ws.cell(row=row, column=6, value=0); estilo(c, "entrada"); c.number_format = 'R$ #,##0.00'
        c = ws.cell(row=row, column=7)
        c.value = f'=IF(AND(E{row}>0,F{row}>0),E{row}*F{row},0)'; estilo(c, "calculado"); c.number_format = 'R$ #,##0.00'
    ws["B27"].value = "Investimento por ha:"; estilo(ws["B27"], "calculado", negrito=True, tamanho=11)
    ws["B27"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 27, "G", '=SUM(G11:G25)', 'R$ #,##0.00', negrito=True, tamanho=11)
    nomear(wb, "invest_ha", "5.Implantacao", "G27")
    rotulo(ws, 28, "Área implantada (ha):", "B")
    calc(ws, 28, "G", '=qtd_val', "#,##0.00")
    nomear(wb, "area_impl", "5.Implantacao", "G28")
    ws["B29"].value = "INVESTIMENTO TOTAL (ano 0):"; estilo(ws["B29"], "ok", negrito=True, tamanho=12)
    ws["B29"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 29, "G", '=G27*G28', 'R$ #,##0.00', negrito=True, tamanho=12); nomear(wb, "invest_total", "5.Implantacao", "G29")
    ws["B31"].value = "Depreciação anual da formação (automática):"; estilo(ws["B31"], "calculado", negrito=True)
    ws["B31"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 31, "G", '=IF(AND(G28>0,vida_util>0),G29/vida_util,0)', 'R$ #,##0.00', negrito=True)
    nomear(wb, "dep_formacao", "5.Implantacao", "G31")
    auto_largura(ws)

    # ABA 6 — Fixos
    ws = wb.create_sheet("6.Fixos"); ws.sheet_properties.tabColor = "2E5C3F"
    titulo_secao(ws, 1, "6. CUSTOS FIXOS E REMUNERAÇÃO DOS FATORES", "B", "H")
    titulo_secao(ws, 3, "DEPRECIAÇÃO (BENS DE CAPITAL)", "B", "H")
    fix_h = ["Nº", "Descrição", "Aquisição (R$)", "Residual (R$)", "Vida útil (anos)", "Depreciação anual (R$)", "Capital médio (R$)"]
    for c_idx, h in enumerate(fix_h, 2):
        cell = ws.cell(row=4, column=c_idx, value=h); estilo(cell, "titulo", tamanho=10)
    for row in range(5, 15):
        ws.cell(row=row, column=2, value=row-4).alignment = Alignment(horizontal="center", vertical="center")
        estilo(ws.cell(row=row, column=3), "entrada")
        for col in [4,5,6]:
            c = ws.cell(row=row, column=col, value=0); estilo(c, "entrada")
            c.number_format = '0' if col == 6 else 'R$ #,##0.00'
        c = ws.cell(row=row, column=7)
        c.value = f'=IF(AND(D{row}>0,F{row}>0),MAX(0,(D{row}-E{row})/F{row}),0)'; estilo(c, "calculado"); c.number_format = 'R$ #,##0.00'
        c = ws.cell(row=row, column=8)
        c.value = f'=IF(D{row}>0,(D{row}+E{row})/2,0)'; estilo(c, "calculado"); c.number_format = 'R$ #,##0.00'
    ws["B16"].value = "Depreciação bens de capital:"; estilo(ws["B16"], "calculado", negrito=True, tamanho=11)
    ws["B16"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 16, "G", '=SUM(G5:G14)', 'R$ #,##0.00', negrito=True, tamanho=11); nomear(wb, "dep_bens", "6.Fixos", "G16")
    ws["B17"].value = "Depreciação formação (automática):"; estilo(ws["B17"], "calculado", negrito=True)
    ws["B17"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 17, "G", '=dep_formacao', 'R$ #,##0.00', negrito=True)
    ws["B18"].value = "DEPRECIAÇÃO TOTAL:"; estilo(ws["B18"], "ok", negrito=True, tamanho=11)
    ws["B18"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 18, "G", '=G16+G17', 'R$ #,##0.00', negrito=True, tamanho=11); nomear(wb, "dep_total", "6.Fixos", "G18")
    ws["B20"].value = "Capital médio bens:"; estilo(ws["B20"], "calculado", negrito=True)
    ws["B20"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 20, "G", '=SUM(H5:H14)', 'R$ #,##0.00', negrito=True)
    ws["B21"].value = "(+) Capital médio formação (invest/2):"; estilo(ws["B21"], "calculado", negrito=True)
    ws["B21"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 21, "G", '=IF(invest_total>0,invest_total/2,0)', 'R$ #,##0.00', negrito=True)
    ws["B22"].value = "CAPITAL MÉDIO TOTAL:"; estilo(ws["B22"], "ok", negrito=True, tamanho=11)
    ws["B22"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 22, "G", '=G20+G21', 'R$ #,##0.00', negrito=True, tamanho=11); nomear(wb, "cap_medio", "6.Fixos", "G22")
    titulo_secao(ws, 24, "OUTROS FIXOS E REMUNERAÇÃO", "B", "H")
    for l, v, t, n, f in [(25,"C25","Pró-labore anual (R$):","prol_val",'R$ #,##0.00'),
                            (26,"C26","Outros fixos (R$):","outros_fixos",'R$ #,##0.00'),
                            (27,"C27","Taxa remuneração capital (% a.a.):","taxa_cap",'0.00"%"'),
                            (28,"C28","Arrendamento (R$/ha):","arrend",'R$ #,##0.00')]:
        rotulo(ws, l, t, "B"); c = entrada(ws, l, v, 0, f); nomear(wb, n, "6.Fixos", v)
    rotulo(ws, 29, "Área para arrendamento (ha):", "B")
    calc(ws, 29, "C", '=qtd_val', "#,##0.00")
    nomear(wb, "area_fixos", "6.Fixos", "C29")
    ws["E25"].value = "Remuneração do capital:"; estilo(ws["E25"], "calculado", negrito=True)
    ws["E25"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 25, "G", '=IF(AND(cap_medio>0,taxa_cap>0),cap_medio*(taxa_cap/100),0)', 'R$ #,##0.00', negrito=True); nomear(wb, "rem_cap", "6.Fixos", "G25")
    ws["E26"].value = "Remuneração da terra:"; estilo(ws["E26"], "calculado", negrito=True)
    ws["E26"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 26, "G", '=IF(AND(arrend>0,area_fixos>0),arrend*area_fixos,0)', 'R$ #,##0.00', negrito=True); nomear(wb, "rem_terra", "6.Fixos", "G26")
    for l, t, v, f, d in [(31,"COE (do passo 4):","G31","=coe_total",False),
                            (32,"(+) Depreciação total:","G32","=dep_total",False),
                            (33,"(+) Pró-labore:","G33","=prol_val",False),
                            (34,"(+) Outros fixos:","G34","=outros_fixos",False),
                            (35,"COT:","G35","=G31+G32+G33+G34",True),
                            (36,"(+) Remuneração do capital:","G36","=rem_cap",False),
                            (37,"(+) Remuneração da terra:","G37","=rem_terra",False),
                            (38,"CT (CUSTO TOTAL):","G38","=G35+G36+G37",True)]:
        cl = ws[f"B{l}"]; cl.value = t; estilo(cl, "calculado", negrito=True); cl.alignment = Alignment(horizontal="right", vertical="center")
        if d:
            calc(ws, int(v[1:]), v[0], f, 'R$ #,##0.00', negrito=True, tamanho=12 if "CT" in t else 11)
        else:
            calc(ws, int(v[1:]), v[0], f, 'R$ #,##0.00')
    nomear(wb, "cot_val", "6.Fixos", "G35"); nomear(wb, "ct_val", "6.Fixos", "G38")
    auto_largura(ws)

    # ABA 7 — Resultado
    ws = wb.create_sheet("7.Resultado"); ws.sheet_properties.tabColor = "A8552E"
    titulo_secao(ws, 1, "7. RESULTADO E INDICADORES", "B", "H")
    ws["B3"].value = "RESULTADO POR NÍVEL"; estilo(ws["B3"], "titulo", tamanho=12); ws.merge_cells("B3:E3")
    dre = [("B4","Receita Líquida (RL):","=rl_val",False),("B5","(−) COE:","=coe_total",False),
           ("B6","Margem Bruta:","=C4-C5",True),("B7","(−) Fixos (COT − COE):","=cot_val-coe_total",False),
           ("B8","Margem Líquida:","=C6-C7",True),("B9","(−) Fatores (CT − COT):","=ct_val-cot_val",False),
           ("B10","RESULTADO:","=C8-C9",True)]
    for lbl, txt, formula, destaque in dre:
        ws[lbl] = txt; estilo(ws[lbl], "calculado" if not destaque else "ok", negrito=True)
        ws[lbl].alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=int(lbl[1:]), column=3); c.value = formula
        if destaque: estilo(c, "ok", negrito=True, tamanho=12 if "RESULTADO" in txt else 11)
        else: estilo(c, "calculado", negrito=True)
        c.number_format = 'R$ #,##0.00'
    nomear(wb, "margem_bruta", "7.Resultado", "C6"); nomear(wb, "margem_liq", "7.Resultado", "C8"); nomear(wb, "resultado_final", "7.Resultado", "C10")
    ws["B12"].value = "INDICADORES ECONÔMICOS"; estilo(ws["B12"], "titulo", tamanho=12); ws.merge_cells("B12:E12")
    ind = [("B13","Produtividade de nivelamento:",'=IF(AND(qtd_val>0,preco_val>0),coe_total/(qtd_val*preco_val),0)',"#,##0.00"),
           ("B14","Preço de nivelamento:",'=IF(AND(qtd_val>0,prod_val>0),coe_total/(qtd_val*prod_val),0)','R$ #,##0.00'),
           ("B15","Margem de segurança:",'=IF(AND(prod_val>0,B13>0),(prod_val-B13)/prod_val,0)','0.00%'),
           ("B16","Lucratividade:",'=IF(rb_val>0,resultado_final/rb_val,0)','0.00%'),
           ("B17","Rentabilidade (ROI):",'=IF(coe_total>0,resultado_final/coe_total,0)','0.00%'),
           ("B18","Eficiência econômica (RB/CT):",'=IF(ct_val>0,rb_val/ct_val,0)','0.00'),
           ("B19","Ponto de equilíbrio (unidades):",'=IF(rl_val>0,ct_val/(rl_val/qtd_val),0)','#,##0.00')]
    for lbl, txt, formula, fmt in ind:
        ws[lbl] = txt; estilo(ws[lbl], "calculado", negrito=True); ws[lbl].alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=int(lbl[1:]), column=3); c.value = formula; estilo(c, "calculado", negrito=True); c.number_format = fmt
    ws["B21"].value = "AVALIAÇÃO PLURIANUAL (PERENES)"; estilo(ws["B21"], "titulo", tamanho=12); ws.merge_cells("B21:H21")
    rotulo(ws, 22, "Custo de formação anual TOTAL (R$):", "B"); entrada(ws, 22, "C", 0, 'R$ #,##0.00'); nomear(wb, "custo_form", "7.Resultado", "C22")
    rotulo(ws, 23, "Taxa de desconto (% a.a.):", "B"); entrada(ws, 23, "C", 6, '0.00"%"'); nomear(wb, "taxa_desc", "7.Resultado", "C23")
    fluxo_h = ["Ano", "Fase", "Fluxo (R$)", "Descontado (R$)", "Acumulado (R$)"]
    for c_idx, h in enumerate(fluxo_h, 2):
        cell = ws.cell(row=25, column=c_idx, value=h); estilo(cell, "titulo", tamanho=10)
    for row in range(26, 66):
        ano = row - 26
        ws.cell(row=row, column=2, value=ano).alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(row=row, column=3)
        # CORREÇÃO: fase para de ser "Produção" após form_anos+vida_util (antes corria os 40 anos inteiros)
        c.value = f'=IF(B{row}=0,"Implantação",IF(B{row}<=form_anos,"Formação",IF(B{row}<=form_anos+vida_util,"Produção","")))'
        estilo(c, "calculado")
        c = ws.cell(row=row, column=4)
        # CORREÇÃO CRÍTICO-001: fluxo exclui TODA depreciação (dep_total).
        # CORREÇÃO: fluxo de produção zera após form_anos+vida_util, em vez de continuar até a linha 65.
        c.value = f'=IF(B{row}=0,-invest_total,IF(B{row}<=form_anos,-custo_form,IF(B{row}<=form_anos+vida_util,rl_val-(ct_val-dep_total),0)))'
        estilo(c, "calculado"); c.number_format = 'R$ #,##0.00'
        c = ws.cell(row=row, column=5)
        c.value = f'=IF(D{row}<>"",D{row}/(1+taxa_desc/100)^B{row},0)'; estilo(c, "calculado"); c.number_format = 'R$ #,##0.00'
        c = ws.cell(row=row, column=6)
        c.value = f'=E{row}' if row == 26 else f'=F{row-1}+E{row}'; estilo(c, "calculado"); c.number_format = 'R$ #,##0.00'
    ws["B68"].value = "VPL:"; estilo(ws["B68"], "ok", negrito=True, tamanho=12); ws["B68"].alignment = Alignment(horizontal="right", vertical="center")
    # CORREÇÃO: NPV() já desconta cada valor; usar a coluna D (fluxo bruto), não a E (já descontada),
    # senão cada ano é descontado duas vezes (erro composto: até (1+i)^78 em vez de (1+i)^39 no último ano).
    calc(ws, 68, "C", '=D26+NPV(taxa_desc/100,D27:D65)', 'R$ #,##0.00', negrito=True, tamanho=12); nomear(wb, "vpl_val", "7.Resultado", "C68")
    ws["B69"].value = "TIR:"; estilo(ws["B69"], "ok", negrito=True, tamanho=12); ws["B69"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 69, "C", '=IF(ISERROR(IRR(D26:D65)),"n/d",IRR(D26:D65))', '0.00%', negrito=True, tamanho=12); nomear(wb, "tir_val", "7.Resultado", "C69")
    ws["B70"].value = "Payback (anos):"; estilo(ws["B70"], "ok", negrito=True, tamanho=12); ws["B70"].alignment = Alignment(horizontal="right", vertical="center")
    calc(ws, 70, "C", '=IF(COUNTIF(F26:F65,">=0")=0,"n/d",MATCH(TRUE,F26:F65>=0,0)-1+ABS(INDEX(F26:F65,MATCH(TRUE,F26:F65>=0,0)-1))/INDEX(E26:E65,MATCH(TRUE,F26:F65>=0,0)))', '0.0', negrito=True, tamanho=12)
    ws["B72"].value = "NOTA: Fluxo de caixa exclui TODA depreciação (correção CRÍTICO-001)."
    ws["B72"].font = Font(italic=True, color="B71C1C", size=9); ws.merge_cells("B72:H72")
    auto_largura(ws)

    # ABA 8 — Relatório
    ws = wb.create_sheet("8.Relatorio"); ws.sheet_properties.tabColor = "3D4A2A"
    ws.merge_cells("B1:H1"); ws["B1"].value = "PROJETO DE VIABILIDADE AGROPECUÁRIA"
    estilo(ws["B1"], "titulo", tamanho=16); ws.row_dimensions[1].height = 30
    ws.merge_cells("B2:H2"); ws["B2"].value = '=projeto_val'; estilo(ws["B2"], "destaque", negrito=True, tamanho=13); ws.row_dimensions[2].height = 24
    meta = [("B4",'="Produtor: " & produtor_val'),("B5",'="Unidade: " & unidade_val'),
            ("B6",'="Município: " & municipio_val'),("B7",'="Data: " & data_val'),
            ("B8",'="Template: " & VLOOKUP(template_val,Ref!$A$1:$K$17,2,FALSE)'),
            ("B9",'="Escala: " & TEXT(qtd_val,"#,##0.00") & " " & VLOOKUP(template_val,Ref!$A$1:$K$17,3,FALSE)')]
    for cel, formula in meta:
        ws[cel].value = formula; ws[cel].font = Font(size=10, name="Calibri")
        ws[cel].alignment = Alignment(horizontal="left", vertical="center")
    ws["B11"].value = "DEMONSTRAÇÃO DE RESULTADO"; estilo(ws["B11"], "titulo", tamanho=12); ws.merge_cells("B11:D11")
    dre_rel = [("B12","Receita Bruta:","=rb_val"),("B13","(−) Comercialização:","=comerc_rs"),
               ("B14","(−) Funrural:","=funrural_rs"),("B15","Receita Líquida:","=rl_val"),
               ("B16","(−) COE:","=coe_total"),("B17","Margem Bruta:","=margem_bruta"),
               ("B18","(−) Fixos:","=cot_val-coe_total"),("B19","Margem Líquida:","=margem_liq"),
               ("B20","(−) Fatores:","=ct_val-cot_val"),("B21","RESULTADO:","=resultado_final")]
    for lbl, txt, formula in dre_rel:
        ws[lbl] = txt; estilo(ws[lbl], "calculado", negrito=True); ws[lbl].alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=int(lbl[1:]), column=3); c.value = formula; estilo(c, "calculado", negrito=True)
        c.number_format = 'R$ #,##0.00'
        if "RESULTADO" in txt: estilo(c, "ok", negrito=True, tamanho=12); c.number_format = 'R$ #,##0.00'
    ws["B23"].value = "INDICADORES"; estilo(ws["B23"], "titulo", tamanho=12); ws.merge_cells("B23:D23")
    ind_rel = [("B24","Eficiência econômica (RB/CT):",'=IF(ct_val>0,rb_val/ct_val,0)','0.00'),
               ("B25","Lucratividade:",'=IF(rb_val>0,resultado_final/rb_val,0)','0.00%'),
               ("B26","Rentabilidade (ROI):",'=IF(coe_total>0,resultado_final/coe_total,0)','0.00%'),
               ("B27","Margem de segurança:",'=IF(AND(prod_val>0,B13>0),(prod_val-B13)/prod_val,0)','0.00%')]
    for lbl, txt, formula, fmt in ind_rel:
        ws[lbl] = txt; estilo(ws[lbl], "calculado", negrito=True); ws[lbl].alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=int(lbl[1:]), column=3); c.value = formula; estilo(c, "calculado", negrito=True); c.number_format = fmt
    ws["B29"].value = "AVALIAÇÃO PLURIANUAL (se perene)"; estilo(ws["B29"], "titulo", tamanho=12); ws.merge_cells("B29:D29")
    perene_rel = [("B30","Investimento total:","=invest_total",'R$ #,##0.00'),("B31","VPL:","=vpl_val",'R$ #,##0.00'),
                  ("B32","TIR:","=tir_val",'0.00%'),("B33","Cultura:","=cultura_val","General"),
                  ("B34","Anos formação:","=form_anos","0"),("B35","Vida útil:","=vida_util","0")]
    for lbl, txt, formula, fmt in perene_rel:
        ws[lbl] = txt; estilo(ws[lbl], "calculado", negrito=True); ws[lbl].alignment = Alignment(horizontal="right", vertical="center")
        c = ws.cell(row=int(lbl[1:]), column=3); c.value = formula; estilo(c, "calculado", negrito=True); c.number_format = fmt
    ws["B37"].value = "Metodologia: CNA/Campo Futuro. COE → COT → CT. Números ilustrativos."
    ws["B37"].font = Font(italic=True, color="666666", size=9); ws.merge_cells("B37:H37")
    ws["B38"].value = "NOTAS: (1) Fluxo de caixa (aba 7) exclui toda depreciação e para no fim da vida útil produtiva. (2) Área de implantação e de arrendamento herdam a área do passo 2 (Receita). (3) Ciclos por ano e mortalidade de atividades de giro não são modelados nesta planilha."
    ws["B38"].font = Font(italic=True, color="B71C1C", size=9); ws.merge_cells("B38:H38")
    auto_largura(ws)

    # Proteger planilhas (células calculadas bloqueadas)
    for ws_name in wb.sheetnames:
        if ws_name != "Ref":
            wb[ws_name].protection.sheet = True
    return wb

if __name__ == "__main__":
    wb = construir_workbook()
    wb.save(OUTPUT_FILE)
    print(f"[OK] Arquivo gerado: {OUTPUT_FILE}")
    print(f"  Abas: {wb.sheetnames}")
